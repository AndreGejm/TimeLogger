import os
import sys
from datetime import datetime, timedelta
import openpyxl

# Configuration DECLARE your path here
primary_excel_path = "/Timelog.xlsx"
backup_folder_path = "/Backup"
event_record_path = "/event_record.txt"

def create_or_load_workbook(file_path):
    """Load an existing Excel workbook or create a new one if it doesn't exist."""
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Total Logged Time'
        sheet['C1'] = 'Weekly Total'
        sheet['F1'] = 'Errors/Warnings'
    return workbook

def save_backup(workbook):
    """Create a backup of the workbook in the specified backup directory."""
    current_week = datetime.now().isocalendar()[1]
    backup_file_name = f"Timelog (Week {current_week}).xlsx"
    backup_file_path = os.path.join(backup_folder_path, backup_file_name)
    if not os.path.exists(backup_folder_path):
        os.makedirs(backup_folder_path)
    workbook.save(backup_file_path)

def record_event(event_type):
    """Record a logon or logoff event with a timestamp."""
    now = datetime.now()
    with open(event_record_path, "a") as file:
        file.write(f"{event_type},{now.isoformat()}\n")

def parse_timedelta(time_str):
    """Safely parse a HH:MM:SS time string into a timedelta object."""
    try:
        hours, minutes, seconds = map(int, time_str.split(":"))
        return timedelta(hours=hours, minutes=minutes, seconds=seconds)
    except (ValueError, AttributeError):
        return timedelta()

def find_or_create_today_row(sheet, date_str):
    """Find the row for the given date or create a new row if it doesn't exist."""
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        if cell_value and isinstance(cell_value, datetime) and cell_value.date().isoformat() == date_str:
            return row
    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1).value = datetime.strptime(date_str, "%Y-%m-%d")
    return new_row
def calculate_logged_time():
    """
    Calculate the total logged time for each day based on all logon and logoff events.
    Replace the data in column B with the total calculated time from the event file for that date.
    """
    if not os.path.exists(event_record_path):
        print("Event record file does not exist.")
        return

    workbook = create_or_load_workbook(primary_excel_path)
    sheet = workbook.active

    # Initialize a dictionary to accumulate total time for each date
    total_time_per_date = {}

    with open(event_record_path, "r") as file:
        lines = file.readlines()

    for line in lines:
        line = line.strip()
        if line:
            event_type, timestamp = line.split(',')
            event_datetime = datetime.fromisoformat(timestamp)
            date_str = event_datetime.date().isoformat()

            # Initialize the date key in the dictionary if not already present
            if date_str not in total_time_per_date:
                total_time_per_date[date_str] = {'logon': [], 'logoff': []}
            
            # Append datetime object to respective list
            if event_type in ['logon', 'logoff']:
                total_time_per_date[date_str][event_type].append(event_datetime)

    # Process each date to calculate the total logged time
    for date_str, events in total_time_per_date.items():
        logon_times = sorted(events['logon'])
        logoff_times = sorted(events['logoff'])
        total_time = timedelta()

        # Assuming each logon is paired with the next chronological logoff
        for logon_time in logon_times:
            for logoff_time in logoff_times:
                if logoff_time > logon_time:
                    total_time += logoff_time - logon_time
                    logoff_times.remove(logoff_time)
                    break

        # Find or create the row for the current date
        row = find_or_create_today_row(sheet, date_str)
        
        # Convert total time to a formatted string and update Column B
        hours, remainder = divmod(int(total_time.total_seconds()), 3600)
        minutes, seconds = divmod(remainder, 60)
        sheet.cell(row=row, column=2).value = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    workbook.save(primary_excel_path)


def main(event_type):
    """Main function to handle logon or logoff events."""
    if event_type not in ["logon", "logoff"]:
        print("Invalid event type. Please specify 'logon' or 'logoff'.")
        return
    record_event(event_type)
    if event_type == "logoff":
        calculate_logged_time()
        workbook = create_or_load_workbook(primary_excel_path)
        save_backup(workbook)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        main(sys.argv[1])
    else:
        print("Please specify the event type ('logon' or 'logoff').")
